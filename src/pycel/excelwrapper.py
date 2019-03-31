"""
    ExcelComWrapper : Must be run on Windows as it requires a COM link
                      to an Excel instance.
    ExcelOpxWrapper : Can be run anywhere but only with post 2010 Excel formats
"""

import abc
import collections
import os
from unittest import mock

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.cell.read_only import EMPTY_CELL
from openpyxl.utils import datetime as opxl_dt
from pycel.excelutil import AddressCell, AddressRange, flatten, MAX_ROW

ARRAY_FORMULA_NAME = '=CSE_INDEX'
ARRAY_FORMULA_FORMAT = '{}(%s,%s,%s,%s,%s)'.format(ARRAY_FORMULA_NAME)


class ExcelWrapper:
    __metaclass__ = abc.ABCMeta

    RangeData = collections.namedtuple('RangeData', 'address formula values')

    @abc.abstractmethod
    def connect(self):
        """"""

    @abc.abstractmethod
    def get_range(self, address):
        """"""

    @abc.abstractmethod
    def get_used_range(self):
        """"""

    @abc.abstractmethod
    def get_active_sheet_name(self):
        """"""

    def get_formula_from_range(self, address):
        if not isinstance(address, (AddressRange, AddressCell)):
            address = AddressRange(address)
        result = self.get_range(address)
        if isinstance(address, AddressCell):
            return result.formula if result.formula.startswith("=") else None
        else:
            return tuple(tuple(
                self.get_formula_from_range(a) for a in row
            ) for row in result.resolve_range)

    def get_formula_or_value(self, address):
        if not isinstance(address, (AddressRange, AddressCell)):
            address = AddressRange(address)
        result = self.get_range(address)
        if isinstance(address, AddressCell):
            return result.formula or result.values
        else:
            return tuple(tuple(
                self.get_formula_or_value(a) for a in row
            ) for row in result.resolve_range)


class _OpxRange(ExcelWrapper.RangeData):
    """ Excel range wrapper that distributes reduced api used by compiler
        (Formula & Value)
    """
    def __new__(cls, cells, cells_dataonly, address):
        formula = None
        value = cells[0][0].value
        if isinstance(value, str) and value.startswith(ARRAY_FORMULA_NAME):
            # if this range refers to a CSE Array Formula, get the formula
            front = cells[0][0].value.rsplit(',', 4)[0]
            if all(c.value and c.value.startswith(front)
                   for c in flatten(cells)):
                formula = '={%s}' % front[len(ARRAY_FORMULA_NAME) + 1:]

        values = tuple(tuple(cell.value for cell in row)
                       for row in cells_dataonly)
        return ExcelWrapper.RangeData.__new__(cls, address, formula, values)

    @classmethod
    def cell_to_formula(cls, cell):
        if cell.value is None:
            return ''
        else:
            formula = str(cell.value)
            if not formula.startswith('='):
                return ''

            elif formula.startswith(ARRAY_FORMULA_NAME):
                params = formula[len(ARRAY_FORMULA_NAME) + 1:-1].rsplit(',', 4)
                start_row = cell.row - int(params[1]) + 1
                start_col_idx = cell.col_idx - int(params[2]) + 1
                end_row = start_row + int(params[3]) - 1
                end_col_idx = start_col_idx + int(params[4]) - 1
                cse_range = AddressRange(
                    (start_col_idx, start_row, end_col_idx, end_row),
                    sheet=cell.parent.title)
                return '=index({},{},{})'.format(
                    cse_range.quoted_address, *params[1:3])
            else:
                return formula

    @property
    def resolve_range(self):
        return AddressRange(
            (self.address.start.col_idx,
             self.address.start.row,
             self.address.start.col_idx + len(self.values[0]) - 1,
             self.address.start.row + len(self.values) - 1),
            sheet=self.address.sheet
        ).resolve_range


class _OpxCell(_OpxRange):
    """ Excel cell wrapper that distributes reduced api used by compiler
        (Formula & Value)
    """
    def __new__(cls, cell, cell_dataonly, address):
        assert isinstance(address, AddressCell)
        return ExcelWrapper.RangeData.__new__(
            cls, address, cls.cell_to_formula(cell), cell_dataonly.value)


class ExcelOpxWrapper(ExcelWrapper):
    """ OpenPyXl implementation for ExcelWrapper interface """

    def __init__(self, filename, app=None):
        super(ExcelWrapper, self).__init__()

        self.filename = os.path.abspath(filename)
        self._defined_names = None
        self._tables = None
        self._table_refs = {}
        self.workbook = None
        self.workbook_dataonly = None

    @property
    def defined_names(self):
        if self.workbook is not None and self._defined_names is None:
            self._defined_names = {}

            for defined_name in self.workbook.defined_names.definedName:
                for worksheet, range_alias in defined_name.destinations:
                    if worksheet in self.workbook:
                        self._defined_names[str(defined_name.name)] = (
                            range_alias, worksheet)
        return self._defined_names

    def table(self, table_name):
        """ Return the table and the sheet it was found on

        :param table_name: name of table to retrieve
        :return: table, sheet_name
        """
        # table names are case insensitive
        if self._tables is None:
            TableAndSheet = collections.namedtuple(
                'TableAndSheet', 'table, sheet_name')
            self._tables = {
                t.name.lower(): TableAndSheet(t, ws.title)
                for ws in self.workbook for t in ws._tables}
            self._tables[None] = TableAndSheet(None, None)
        return self._tables.get(table_name.lower(), self._tables[None])

    def table_name_containing(self, address):
        """ Return the table name containing the address given """
        address = AddressCell(address)
        if address not in self._table_refs:
            for t in self.workbook[address.sheet]._tables:
                if address in AddressRange(t.ref):
                    self._table_refs[address] = t.name.lower()
                    break

        return self._table_refs.get(address)

    def connect(self):
        self.workbook = load_workbook(self.filename)
        self.workbook_dataonly = load_workbook(
            self.filename, data_only=True, read_only=True)

        # expand array formulas
        for ws in self.workbook:
            for address, props in ws.formula_attributes.items():
                if props.get('t') != 'array':
                    continue  # pragma: no cover

                # get the reference address for the array formula
                ref_addr = AddressRange(props.get('ref'))

                if isinstance(ref_addr, AddressRange):
                    formula = ws[address].value
                    for i, row in enumerate(ref_addr.rows, start=1):
                        for j, addr in enumerate(row, start=1):
                            ws[addr.coordinate] = ARRAY_FORMULA_FORMAT % (
                                formula[1:], i, j, *ref_addr.size)

        # ::HACK:: this is only needed because openpyxl does not define
        # iter_cols for read only workbooks
        def _iter_cols(self, min_col=None, max_col=None, min_row=None,
                       max_row=None, values_only=False):
            # In the case of lookup for something like C:D, openpyxl
            # attempts to use iter_cols() which is not defined for read_only
            yield from zip(*self.iter_rows(min_col=min_col, max_col=max_col))

        import types
        for sheet in self.workbook_dataonly:
            sheet.iter_cols = types.MethodType(_iter_cols, sheet)

    def set_sheet(self, s):
        self.workbook.active = self.workbook.index(self.workbook[s])
        self.workbook_dataonly.active = self.workbook_dataonly.index(
            self.workbook_dataonly[s])
        return self.workbook.active

    @staticmethod
    def from_excel(value, offset=opxl_dt.CALENDAR_WINDOWS_1900):
        # ::HACK:: excel thinks that 1900/02/29 was a thing.  In certain
        # circumstances openpyxl will return a datetime.  This is a problem
        # as we don't want them, and having been mapped to datetime
        # information may have been lost, so ignore the conversions
        return value

    def get_range(self, address):
        if not isinstance(address, (AddressRange, AddressCell)):
            address = AddressRange(address)

        if address.has_sheet:
            sheet = self.workbook[address.sheet]
            sheet_dataonly = self.workbook_dataonly[address.sheet]
        else:
            sheet = self.workbook.active
            sheet_dataonly = self.workbook_dataonly.active

        with mock.patch('openpyxl.worksheet._reader.from_excel',
                        self.from_excel):
            # work around type coercion to datetime that causes some issues

            cells = sheet[address.coordinate]
            if isinstance(cells, Cell):
                cell = cells
                cell_dataonly = sheet_dataonly[address.coordinate]
                return _OpxCell(cell, cell_dataonly, address)

            else:
                cells_dataonly = sheet_dataonly[address.coordinate]
                addr_size = address.size

                if 1 in addr_size:
                    if cells_dataonly \
                            and not isinstance(cells_dataonly[0], tuple):
                        # openpyxl returns a one dimensional structure for some
                        if addr_size.width == 1:
                            cells = tuple((c,) for c in cells)
                            cells_dataonly = tuple(
                                (c,) for c in cells_dataonly)
                        else:
                            cells = (cells,)
                            cells_dataonly = (cells_dataonly,)

                elif addr_size.height == MAX_ROW:
                    # openpyxl does iter_cols, we need to transpose
                    cells = tuple(zip(*cells))
                    cells_dataonly = tuple(zip(*cells_dataonly))

                if len(cells) != len(cells_dataonly):
                    # The read_only version of openpyxl worksheet has the
                    # somewhat annoying property of not giving empty rows at the
                    # end.  Which is not the same behavior as the non-readonly
                    # version.  So we need to align the data here by adding
                    # empty rows.
                    empty_row = (EMPTY_CELL, ) * len(cells[0])
                    empty_rows = (empty_row, ) * (
                        len(cells) - len(cells_dataonly))
                    cells_dataonly += empty_rows

                # full range column or row addresses, trim the address
                if len(cells) < addr_size.height or \
                        len(cells[0]) < addr_size.width:
                    start_col = address.start.column or 'A'
                    start_row = address.start.row or 1
                    start_addr = AddressCell(
                        start_col + str(start_row), sheet=address.sheet)

                    stop_addr = start_addr.address_at_offset(
                        len(cells) - 1, len(cells[0]) - 1)

                    address = AddressRange((
                        start_addr.col_idx, start_addr.row,
                        stop_addr.col_idx, stop_addr.row),
                        sheet=address.sheet)

                return _OpxRange(cells, cells_dataonly, address)

    def get_used_range(self):
        return self.workbook.active.iter_rows()

    def get_active_sheet_name(self):
        return self.workbook.active.title
